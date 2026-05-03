import streamlit as st
import pandas as pd
import sqlite3
import os
import re
from datetime import datetime
from PIL import Image
import io
import base64
import matplotlib.pyplot as plt
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image as XlImage

# -----------------------------
# CONFIGURAÇÃO DA PÁGINA
# -----------------------------
st.set_page_config(
    page_title="AILSON PERSONAL TRAINNER",
    page_icon="💪",
    layout="wide",
    initial_sidebar_state="expanded"
)

AZUL = "#1E3A8A"
PRETO = "#111111"
BRANCO = "#FFFFFF"

st.markdown(f"""
<style>
    .stApp {{
        background-color: {PRETO};
    }}
    .css-1d391kg {{
        background-color: {AZUL};
    }}
    .stButton>button {{
        background-color: {AZUL};
        color: {BRANCO};
        border-radius: 8px;
    }}
    .stButton>button:hover {{
        background-color: #2E4A9A;
    }}
    h1, h2, h3, h4, h5, h6, p, label {{
        color: {BRANCO};
    }}
    .stSidebar {{
        background-color: {PRETO};
        color: {BRANCO};
    }}
    .stSidebar .stSelectbox label, .stSidebar .stTextInput label {{
        color: {BRANCO} !important;
    }}
</style>
""", unsafe_allow_html=True)

LOGO_PATH = "logo.png"
if os.path.exists(LOGO_PATH):
    st.sidebar.image(LOGO_PATH, width=180)
else:
    st.sidebar.title("AILSON PERSONAL")
    st.sidebar.markdown("*Trainner*")

# -----------------------------
# BANCO DE DADOS
# -----------------------------
def init_db():
    conn = sqlite3.connect('clientes.db')
    c = conn.cursor()
    c.execute('''CREATE TABLE IF NOT EXISTS clientes (
                    id INTEGER PRIMARY KEY,
                    nome TEXT,
                    sexo TEXT DEFAULT "Masculino",
                    idade INTEGER,
                    peso REAL,
                    altura REAL,
                    percentual_gordura REAL,
                    nivel TEXT,
                    objetivo TEXT,
                    modalidade TEXT DEFAULT "Geral",
                    agachamento_1rm REAL,
                    supino_1rm REAL,
                    terra_1rm REAL,
                    pegada_direita REAL,
                    pegada_esquerda REAL,
                    historico TEXT DEFAULT ""
                )''')
    c.execute('''CREATE TABLE IF NOT EXISTS fotos (
                    id INTEGER PRIMARY KEY,
                    cliente_id INTEGER,
                    data TEXT,
                    foto_frente BLOB,
                    foto_costas BLOB,
                    foto_perfil BLOB
                )''')
    c.execute('''CREATE TABLE IF NOT EXISTS avaliacao_postural (
                    id INTEGER PRIMARY KEY,
                    cliente_id INTEGER,
                    data TEXT,
                    cabeca TEXT DEFAULT "Normal",
                    ombros TEXT DEFAULT "Normal",
                    coluna TEXT DEFAULT "Normal",
                    quadril TEXT DEFAULT "Normal",
                    joelhos TEXT DEFAULT "Normal",
                    pes TEXT DEFAULT "Normal"
                )''')
    c.execute('''CREATE TABLE IF NOT EXISTS treinos_realizados (
                    id INTEGER PRIMARY KEY,
                    cliente_id INTEGER,
                    data TEXT,
                    exercicio TEXT,
                    series INTEGER,
                    repeticoes INTEGER,
                    carga REAL,
                    observacoes TEXT
                )''')
    c.execute('''CREATE TABLE IF NOT EXISTS alimentos (
                    id INTEGER PRIMARY KEY,
                    nome TEXT,
                    calorias REAL,
                    proteinas REAL,
                    carboidratos REAL,
                    gorduras REAL,
                    porcao REAL
                )''')
    c.execute("SELECT COUNT(*) FROM alimentos")
    if c.fetchone()[0] == 0:
        alimentos_padrao = [
            ("Arroz integral cozido", 123, 2.6, 25.8, 1.0, 100),
            ("Feijão carioca cozido", 76, 4.8, 13.6, 0.5, 100),
            ("Frango grelhado (peito)", 165, 31, 0, 3.6, 100),
            ("Ovo cozido", 155, 12.6, 0.6, 10.6, 100),
            ("Batata doce cozida", 86, 1.6, 20.1, 0.1, 100),
            ("Brócolis cozido", 35, 2.4, 7.2, 0.4, 100),
            ("Banana prata", 98, 1.3, 26, 0.1, 100),
            ("Mamão papaia", 45, 0.5, 11.3, 0.1, 100),
            ("Leite integral", 64, 3.2, 4.9, 3.4, 100),
            ("Iogurte natural", 61, 5.2, 6.6, 1.5, 100),
            ("Aveia em flocos", 389, 13.2, 66.6, 7.2, 100),
            ("Pão integral", 253, 9.4, 49.9, 3.3, 100),
            ("Queijo minas frescal", 264, 17.8, 0.7, 20.9, 100),
            ("Azeite de oliva", 884, 0, 0, 100, 100),
            ("Castanha do pará", 656, 14.3, 12.3, 66.4, 100),
            ("Maçã", 52, 0.3, 13.8, 0.2, 100),
            ("Tomate", 18, 0.9, 3.9, 0.2, 100),
            ("Alface", 15, 1.2, 2.8, 0.2, 100),
            ("Bife de alcatra grelhado", 278, 49.6, 0, 8.5, 100),
            ("Salmão grelhado", 208, 25.3, 0, 12.9, 100),
            ("Atum em água", 116, 25.5, 0, 0.8, 100),
            ("Batata inglesa cozida", 77, 2.0, 18.4, 0.1, 100),
            ("Mandioca cozida", 125, 1.6, 30.1, 0.2, 100),
            ("Clara de ovo", 52, 10.9, 0.7, 0.2, 100),
        ]
        c.executemany("INSERT INTO alimentos (nome, calorias, proteinas, carboidratos, gorduras, porcao) VALUES (?,?,?,?,?,?)", alimentos_padrao)
    conn.commit()
    conn.close()

init_db()

# -----------------------------
# FUNÇÕES AUXILIARES
# -----------------------------
def salvar_cliente(nome, sexo, idade, peso, altura, perc_gord, nivel, objetivo, modalidade, agach, sup, terra, peg_dir, peg_esq):
    conn = sqlite3.connect('clientes.db')
    c = conn.cursor()
    c.execute("INSERT INTO clientes (nome,sexo,idade,peso,altura,percentual_gordura,nivel,objetivo,modalidade,agachamento_1rm,supino_1rm,terra_1rm,pegada_direita,pegada_esquerda) VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?)",
              (nome, sexo, idade, peso, altura, perc_gord, nivel, objetivo, modalidade, agach, sup, terra, peg_dir, peg_esq))
    conn.commit()
    conn.close()

def atualizar_cliente(id_cliente, nome, sexo, idade, peso, altura, perc_gord, nivel, objetivo, modalidade, agach, sup, terra, peg_dir, peg_esq):
    conn = sqlite3.connect('clientes.db')
    c = conn.cursor()
    c.execute("UPDATE clientes SET nome=?, sexo=?, idade=?, peso=?, altura=?, percentual_gordura=?, nivel=?, objetivo=?, modalidade=?, agachamento_1rm=?, supino_1rm=?, terra_1rm=?, pegada_direita=?, pegada_esquerda=? WHERE id=?",
              (nome, sexo, idade, peso, altura, perc_gord, nivel, objetivo, modalidade, agach, sup, terra, peg_dir, peg_esq, id_cliente))
    conn.commit()
    conn.close()

def excluir_cliente(id_cliente):
    conn = sqlite3.connect('clientes.db')
    c = conn.cursor()
    c.execute("DELETE FROM clientes WHERE id=?", (id_cliente,))
    c.execute("DELETE FROM fotos WHERE cliente_id=?", (id_cliente,))
    c.execute("DELETE FROM avaliacao_postural WHERE cliente_id=?", (id_cliente,))
    c.execute("DELETE FROM treinos_realizados WHERE cliente_id=?", (id_cliente,))
    conn.commit()
    conn.close()

def carregar_clientes():
    conn = sqlite3.connect('clientes.db')
    df = pd.read_sql("SELECT id, nome, sexo, idade, peso, altura, percentual_gordura, nivel, objetivo, modalidade FROM clientes", conn)
    conn.close()
    return df

def carregar_cliente(id_cliente):
    conn = sqlite3.connect('clientes.db')
    df = pd.read_sql(f"SELECT * FROM clientes WHERE id={id_cliente}", conn)
    conn.close()
    return df.iloc[0] if not df.empty else None

def salvar_foto(cliente_id, data, frente, costas, perfil):
    conn = sqlite3.connect('clientes.db')
    c = conn.cursor()
    c.execute("INSERT INTO fotos (cliente_id, data, foto_frente, foto_costas, foto_perfil) VALUES (?,?,?,?,?)",
              (cliente_id, data, frente, costas, perfil))
    conn.commit()
    conn.close()

def carregar_fotos(cliente_id):
    conn = sqlite3.connect('clientes.db')
    df = pd.read_sql(f"SELECT * FROM fotos WHERE cliente_id={cliente_id} ORDER BY data DESC", conn)
    conn.close()
    return df

def salvar_avaliacao_postural(cliente_id, data, cabeca, ombros, coluna, quadril, joelhos, pes):
    conn = sqlite3.connect('clientes.db')
    c = conn.cursor()
    c.execute("INSERT INTO avaliacao_postural (cliente_id, data, cabeca, ombros, coluna, quadril, joelhos, pes) VALUES (?,?,?,?,?,?,?)",
              (cliente_id, data, cabeca, ombros, coluna, quadril, joelhos, pes))
    conn.commit()
    conn.close()

def carregar_ultima_postural(cliente_id):
    conn = sqlite3.connect('clientes.db')
    df = pd.read_sql(f"SELECT * FROM avaliacao_postural WHERE cliente_id={cliente_id} ORDER BY data DESC LIMIT 1", conn)
    conn.close()
    return df.iloc[0] if not df.empty else None

def salvar_treino(cliente_id, data, exercicio, series, reps, carga, obs=""):
    conn = sqlite3.connect('clientes.db')
    c = conn.cursor()
    c.execute("INSERT INTO treinos_realizados (cliente_id, data, exercicio, series, repeticoes, carga, observacoes) VALUES (?,?,?,?,?,?,?)",
              (cliente_id, data, exercicio, series, reps, carga, obs))
    conn.commit()
    conn.close()

def carregar_historico_treinos(cliente_id):
    conn = sqlite3.connect('clientes.db')
    df = pd.read_sql(f"SELECT * FROM treinos_realizados WHERE cliente_id={cliente_id} ORDER BY data DESC", conn)
    conn.close()
    return df

# -----------------------------
# FUNÇÕES DE CÁLCULO NUTRICIONAL
# -----------------------------
def calcular_tmb(peso, altura, idade, sexo):
    if sexo == "Masculino":
        return 10 * peso + 6.25 * altura - 5 * idade + 5
    else:
        return 10 * peso + 6.25 * altura - 5 * idade - 161

def calcular_get(tmb, fator_atividade):
    fatores = {
        "Sedentário": 1.2,
        "Pouco ativo": 1.375,
        "Moderadamente ativo": 1.55,
        "Muito ativo": 1.725,
        "Atleta": 1.9
    }
    return tmb * fatores.get(fator_atividade, 1.55)

def distribuir_macros(peso, objetivo, modalidade, get, sexo):
    if objetivo == "Hipertrofia" and modalidade in ["Powerlifting", "Fisiculturismo"]:
        proteina_gkg = 2.2
    elif modalidade in ["Beach Tennis", "Futebol"]:
        proteina_gkg = 1.6
    elif modalidade == "Gestante":
        proteina_gkg = 1.4
    elif objetivo == "Emagrecimento":
        proteina_gkg = 2.0
    else:
        proteina_gkg = 1.8

    proteina_g = peso * proteina_gkg
    calorias_proteina = proteina_g * 4

    if objetivo == "Emagrecimento":
        perc_gord = 0.2
    elif modalidade == "Gestante":
        perc_gord = 0.3
    else:
        perc_gord = 0.25
    calorias_gordura = get * perc_gord
    gordura_g = calorias_gordura / 9

    calorias_carb = get - calorias_proteina - calorias_gordura
    carboidrato_g = calorias_carb / 4

    return round(proteina_g, 1), round(gordura_g, 1), round(carboidrato_g, 1), round(get)

def montar_refeicoes(meta_calorias, meta_prot, meta_carb, meta_gord, num_refeicoes=5):
    conn = sqlite3.connect('clientes.db')
    alimentos_df = pd.read_sql("SELECT * FROM alimentos", conn)
    conn.close()

    distribuicao = [0.20, 0.10, 0.30, 0.15, 0.25]
    refeicoes = []
    for i, perc in enumerate(distribuicao):
        cal = meta_calorias * perc
        prot = meta_prot * perc
        carb = meta_carb * perc
        gord = meta_gord * perc

        sugestoes = []
        if i == 0:
            sugestoes = [("Leite integral", 200), ("Pão integral", 50), ("Mamão papaia", 150)]
        elif i == 1:
            sugestoes = [("Iogurte natural", 200), ("Banana prata", 100)]
        elif i == 2:
            sugestoes = [("Arroz integral cozido", 150), ("Feijão carioca cozido", 100), ("Frango grelhado (peito)", 150), ("Brócolis cozido", 100)]
        elif i == 3:
            sugestoes = [("Aveia em flocos", 30), ("Banana prata", 100)]
        else:
            sugestoes = [("Arroz integral cozido", 100), ("Bife de alcatra grelhado", 120), ("Batata doce cozida", 100), ("Tomate", 50)]

        itens = []
        total_cal = total_prot = total_carb = total_gord = 0
        for nome, gramas in sugestoes:
            alimento = alimentos_df[alimentos_df['nome'] == nome].iloc[0]
            fator = gramas / 100
            cal_item = alimento['calorias'] * fator
            prot_item = alimento['proteinas'] * fator
            carb_item = alimento['carboidratos'] * fator
            gord_item = alimento['gorduras'] * fator
            itens.append((nome, gramas, round(cal_item,1), round(prot_item,1), round(carb_item,1), round(gord_item,1)))
            total_cal += cal_item
            total_prot += prot_item
            total_carb += carb_item
            total_gord += gord_item

        refeicoes.append({
            "nome": f"Refeição {i+1}",
            "itens": itens,
            "total_cal": round(total_cal,1),
            "total_prot": round(total_prot,1),
            "total_carb": round(total_carb,1),
            "total_gord": round(total_gord,1)
        })
    return refeicoes

def exportar_plano_alimentar(cliente, refeicoes, metas):
    wb = Workbook()
    ws = wb.active
    ws.title = "Plano Alimentar"

    if os.path.exists(LOGO_PATH):
        img = XlImage(LOGO_PATH)
        ws.add_image(img, 'A1')

    ws.merge_cells('B2:H2')
    ws['B2'] = "PLANO ALIMENTAR INDIVIDUALIZADO"
    ws['B2'].font = Font(size=14, bold=True)
    ws.merge_cells('B3:H3')
    ws['B3'] = f"Cliente: {cliente['nome']} | Sexo: {cliente['sexo']} | Idade: {cliente['idade']} | Peso: {cliente['peso']} kg"
    ws['B4'] = f"TMB: {metas['tmb']:.0f} kcal | GET: {metas['get']:.0f} kcal | Meta: {metas['meta_cal']:.0f} kcal"
    ws['B5'] = f"Proteína: {metas['prot_g']} g ({metas['cal_prot']:.0f} kcal) | Carboidrato: {metas['carb_g']} g ({metas['cal_carb']:.0f} kcal) | Gordura: {metas['gord_g']} g ({metas['cal_gord']:.0f} kcal)"

    cabeçalhos = ["Refeição", "Alimento", "Gramas", "Calorias", "Proteínas", "Carboidratos", "Gorduras"]
    for c, cab in enumerate(cabeçalhos, 1):
        cell = ws.cell(row=7, column=c, value=cab)
        cell.font = Font(bold=True)

    row = 8
    for ref in refeicoes:
        ws.cell(row, 1, ref['nome']).font = Font(bold=True)
        for item in ref['itens']:
            ws.cell(row, 2, item[0])
            ws.cell(row, 3, item[1])
            ws.cell(row, 4, item[2])
            ws.cell(row, 5, item[3])
            ws.cell(row, 6, item[4])
            ws.cell(row, 7, item[5])
            row += 1
        ws.cell(row, 2, "Total").font = Font(bold=True)
        ws.cell(row, 4, ref['total_cal'])
        ws.cell(row, 5, ref['total_prot'])
        ws.cell(row, 6, ref['total_carb'])
        ws.cell(row, 7, ref['total_gord'])
        row += 2

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output

# -----------------------------
# GERADOR DE PLANILHA ONDULATÓRIA (5 DIAS)
# -----------------------------
def gerar_planilha(cliente, semanas=12, frequencia=3):
    import re
    agach = cliente['agachamento_1rm']
    sup   = cliente['supino_1rm']
    terra = cliente['terra_1rm']
    nivel = cliente['nivel'].split("(")[0].strip()
    objetivo = cliente['objetivo']

    if "Iniciante" in nivel:
        if frequencia == 2:
            rep_schemes = [
                [("4-6 RMs","12-15 RMs")],
                [("8-10 RMs","4-6 RMs")],
                [("12-15 RMs","8-10 RMs")],
                [("4-6 RMs","12-15 RMs")],
                [("8-10 RMs","4-6 RMs")],
                [("12-15 RMs","8-10 RMs")],
                [("4-6 RMs","12-15 RMs")],
                [("8-10 RMs","4-6 RMs")],
                [("12-15 RMs","8-10 RMs")],
                [("4-6 RMs","12-15 RMs")],
                [("8-10 RMs","4-6 RMs")],
                [("12-15 RMs","8-10 RMs")]
            ]
        elif frequencia == 3:
            rep_schemes = [[("4-6 RMs","12-15 RMs","8-10 RMs")]] * 12
        elif frequencia == 4:
            rep_schemes = [[("12-15 RMs","8-10 RMs","12-15 RMs","8-10 RMs")]] * 12
        else:  # 5 dias
            rep_schemes = [[("4-6 RMs","12-15 RMs","8-10 RMs","4-6 RMs","12-15 RMs")]] * 12
    elif "Intermediário" in nivel:
        if frequencia == 2:
            rep_schemes = [
                [("12-15 RMs","18-20 RMs")],
                [("22-25 RMs","12-15 RMs")],
                [("18-20 RMs","22-25 RMs")],
                [("12-15 RMs","18-20 RMs")],
                [("22-25 RMs","12-15 RMs")],
                [("18-20 RMs","22-25 RMs")],
                [("12-15 RMs","18-20 RMs")],
                [("22-25 RMs","12-15 RMs")],
                [("18-20 RMs","22-25 RMs")],
                [("12-15 RMs","18-20 RMs")],
                [("22-25 RMs","12-15 RMs")],
                [("18-20 RMs","22-25 RMs")]
            ]
        elif frequencia == 3:
            rep_schemes = [[("12-15 RMs","18-20 RMs","22-25 RMs")]] * 12
        elif frequencia == 4:
            rep_schemes = [[("12-15 RMs","18-20 RMs","22-25 RMs","12-15 RMs")]] * 12
        else:
            rep_schemes = [[("12-15 RMs","18-20 RMs","22-25 RMs","12-15 RMs","18-20 RMs")]] * 12
    else:  # Avançado / Elite / Competitivo
        if frequencia == 2:
            rep_schemes = [
                [("TA 12-15","TB 8-10")],
                [("TA 4-6","TB 8-10")],
                [("TA 8-10","TB 8-10")],
                [("TA 12-15","TB 8-10")],
                [("TA 12-15","TB 8-10")],
                [("TA 4-6","TB 8-10")],
                [("TA 8-10","TB 8-10")],
                [("TA 12-15","TB 8-10")],
                [("TA 12-15","TB 8-10")],
                [("TA 4-6","TB 8-10")],
                [("TA 8-10","TB 8-10")],
                [("TA 12-15","TB 8-10")]
            ]
        elif frequencia == 3:
            rep_schemes = [[("TA 12-15","TB 8-10","TA 12-15")]] * 12
        elif frequencia == 4:
            rep_schemes = [
                [("TA 12-15","TB 8-10","TA 12-15","TB 8-10")],
                [("TA 4-6","TB 8-10","TA 4-6","TB 8-10")],
                [("TA 8-10","TB 8-10","TA 8-10","TB 8-10")],
                [("TA 12-15","TB 8-10","TA 12-15","TB 8-10")],
                [("TA 12-15","TB 8-10","TA 12-15","TB 8-10")],
                [("TA 4-6","TB 8-10","TA 4-6","TB 8-10")],
                [("TA 8-10","TB 8-10","TA 8-10","TB 8-10")],
                [("TA 12-15","TB 8-10","TA 12-15","TB 8-10")],
                [("TA 12-15","TB 8-10","TA 12-15","TB 8-10")],
                [("TA 4-6","TB 8-10","TA 4-6","TB 8-10")],
                [("TA 8-10","TB 8-10","TA 8-10","TB 8-10")],
                [("TA 12-15","TB 8-10","TA 12-15","TB 8-10")]
            ]
        else:
            rep_schemes = [
                [("TA 12-15","TB 8-10","TA 12-15","TB 8-10","TA 12-15")],
                [("TA 4-6","TB 8-10","TA 4-6","TB 8-10","TA 4-6")],
                [("TA 8-10","TB 8-10","TA 8-10","TB 8-10","TA 8-10")],
                [("TA 12-15","TB 8-10","TA 12-15","TB 8-10","TA 12-15")],
                [("TA 12-15","TB 8-10","TA 12-15","TB 8-10","TA 12-15")],
                [("TA 4-6","TB 8-10","TA 4-6","TB 8-10","TA 4-6")],
                [("TA 8-10","TB 8-10","TA 8-10","TB 8-10","TA 8-10")],
                [("TA 12-15","TB 8-10","TA 12-15","TB 8-10","TA 12-15")],
                [("TA 12-15","TB 8-10","TA 12-15","TB 8-10","TA 12-15")],
                [("TA 4-6","TB 8-10","TA 4-6","TB 8-10","TA 4-6")],
                [("TA 8-10","TB 8-10","TA 8-10","TB 8-10","TA 8-10")],
                [("TA 12-15","TB 8-10","TA 12-15","TB 8-10","TA 12-15")]
            ]

    def exercicios_dia(numero_dia, nivel, modalidade):
        if "Avançado" in nivel or "Elite" in nivel or "Compet" in nivel:
            if numero_dia % 2 == 1:  # TA
                return ["Rosca Biceps", "Rosca Martelo", "Triceps MonoCross", "Triceps Francês", "Desenvolvimento Ombro", "Elevação Lateral"]
            else:
                return ["Agacho Taça", "Push Up", "Remada Curvada", "Agacho Sumo", "Supino Reto", "TRX"]
        else:
            dia1 = ["Agachamento Livre (barra alta)", "Supino Reto", "Remada Curvada"]
            dia2 = ["Levantamento Terra Tradicional", "Desenvolvimento Militar", "Stiff"]
            dia3 = ["Rosca Direta", "Tríceps Testa", "Puxada Alta", "Dips (Paralela)"]
            dia4 = ["Box Squat", "Supino Fechado", "Remada Nórdica"]
            dia5 = ["Good Morning", "Desenvolvimento com Halteres", "Avanço com Barra"]
            if numero_dia == 1: return dia1
            elif numero_dia == 2: return dia2
            elif numero_dia == 3: return dia3
            elif numero_dia == 4: return dia4
            else: return dia5

    dados = []
    for semana in range(1, semanas+1):
        esquema = rep_schemes[semana-1][0]
        for dia in range(1, frequencia+1):
            faixa_rm = esquema[dia-1]
            numeros = re.findall(r'\d+', faixa_rm)
            if len(numeros) >= 2:
                rep_min, rep_max = int(numeros[0]), int(numeros[1])
            else:
                rep_min = rep_max = 10
            series = 3 if "Iniciante" in nivel else 4
            exs = exercicios_dia(dia, nivel, cliente.get('modalidade', 'Geral'))
            for ex in exs[:4]:
                ex_lower = ex.lower()
                if "agachamento" in ex_lower or "box squat" in ex_lower or "agacho" in ex_lower:
                    carga = round(agach * (1 - 0.025*rep_max), 1)
                elif "supino" in ex_lower or "push" in ex_lower:
                    carga = round(sup * (1 - 0.025*rep_max), 1)
                elif "terra" in ex_lower or "stiff" in ex_lower or "good morning" in ex_lower:
                    carga = round(terra * (1 - 0.025*rep_max), 1)
                elif "remada" in ex_lower:
                    carga = round(sup * 0.8 * (1 - 0.025*rep_max), 1)
                elif "desenvolvimento" in ex_lower or "militar" in ex_lower:
                    carga = round(sup * 0.7 * (1 - 0.025*rep_max), 1)
                elif "rosca" in ex_lower or "triceps" in ex_lower or "testa" in ex_lower:
                    carga = round(agach * 0.25 * (1 - 0.025*rep_max), 1)
                elif "puxada" in ex_lower or "dips" in ex_lower or "trx" in ex_lower:
                    carga = 20.0
                else:
                    carga = 50.0
                volume = series * rep_max * carga
                dados.append([semana, f"Semana {semana}", f"Dia {dia}", ex, series, rep_max, carga, volume])

    df = pd.DataFrame(dados, columns=["Semana", "Microciclo", "Dia", "Exercício", "Séries", "Repetições", "Carga (kg)", "Volume Load"])
    return df

def get_table_download_link(df, nome_cliente):
    towrite = io.BytesIO()
    with pd.ExcelWriter(towrite, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name="Treino")
    towrite.seek(0)
    b64 = base64.b64encode(towrite.read()).decode()
    href = f'<a href="data:application/vnd.openxmlformats-officedocument.spreadsheetml.sheet;base64,{b64}" download="treino_{nome_cliente}_{datetime.now().strftime("%Y%m%d")}.xlsx">📥 Baixar Planilha Excel</a>'
    return href

# -----------------------------
# INTERFACE STREAMLIT
# -----------------------------
menu = st.sidebar.selectbox("Menu", [
    "Cadastro de Cliente",
    "Editar / Excluir Clientes",
    "Avaliação & Fotos",
    "Geração de Treino",
    "Plano Alimentar",
    "Histórico & Evolução",
    "Personalizar Exercícios"
])

if menu == "Cadastro de Cliente":
    st.header("➕ Novo Cliente")
    with st.form("form_cliente"):
        col1, col2 = st.columns(2)
        nome = col1.text_input("Nome completo")
        sexo = col2.selectbox("Sexo", ["Masculino", "Feminino"])
        idade = st.number_input("Idade", 6, 100, 30)
        peso = st.number_input("Peso (kg)", 20.0, 200.0, 80.0)
        altura = st.number_input("Altura (cm)", 100.0, 250.0, 170.0)
        perc_gord = st.number_input("Percentual de gordura (%)", 1.0, 60.0, 20.0)
        nivel = st.selectbox("Nível de experiência", [
            "Iniciante (Nível 1)",
            "Básico (Nível 2)",
            "Intermediário (Nível 3)",
            "Avançado (Nível 4)",
            "Elite (Nível 5)",
            "Competitivo (Nível 6)"
        ])
        objetivo = st.selectbox("Objetivo principal", ["Hipertrofia", "Força Máxima", "Potência", "Emagrecimento"])
        modalidade = st.selectbox("Modalidade esportiva", [
            "Geral", "Beach Tennis", "Futebol", "Criança/Adolescente", "Gestante",
            "Powerlifting", "Fisiculturismo", "Musculação Convencional", "V-Taper", "Bikini"
        ])
        st.subheader("Testes de Força (1RM ou Estimado)")
        agach = st.number_input("Agachamento (kg)", 0.0, 500.0, 80.0)
        sup = st.number_input("Supino (kg)", 0.0, 500.0, 60.0)
        terra = st.number_input("Terra (kg)", 0.0, 500.0, 100.0)
        peg_dir = st.number_input("Força de Pegada Mão Direita (kg)", 0.0, 200.0, 40.0)
        peg_esq = st.number_input("Força de Pegada Mão Esquerda (kg)", 0.0, 200.0, 38.0)
        submitted = st.form_submit_button("Salvar Cliente")
        if submitted:
            salvar_cliente(nome, sexo, idade, peso, altura, perc_gord, nivel, objetivo, modalidade, agach, sup, terra, peg_dir, peg_esq)
            st.success(f"Cliente {nome} cadastrado com sucesso!")

elif menu == "Editar / Excluir Clientes":
    st.header("✏️ Editar ou Excluir Clientes")
    clientes_df = carregar_clientes()
    if clientes_df.empty:
        st.warning("Nenhum cliente cadastrado.")
    else:
        cliente_selecionado = st.selectbox("Selecione o cliente", clientes_df['nome'])
        id_cliente = clientes_df[clientes_df['nome'] == cliente_selecionado]['id'].values[0]
        cliente = carregar_cliente(id_cliente)
        tab1, tab2 = st.tabs(["✏️ Editar", "🗑️ Excluir"])
        with tab1:
            with st.form("editar_cliente"):
                nome = st.text_input("Nome completo", cliente['nome'])
                sexo = st.selectbox("Sexo", ["Masculino", "Feminino"], index=0 if cliente['sexo'] == 'Masculino' else 1)
                idade = st.number_input("Idade", 6, 100, cliente['idade'])
                peso = st.number_input("Peso (kg)", 20.0, 200.0, cliente['peso'])
                altura = st.number_input("Altura (cm)", 100.0, 250.0, cliente['altura'])
                perc_gord = st.number_input("Percentual de gordura (%)", 1.0, 60.0, cliente['percentual_gordura'])
                nivel = st.selectbox("Nível de experiência", [
                    "Iniciante (Nível 1)", "Básico (Nível 2)", "Intermediário (Nível 3)",
                    "Avançado (Nível 4)", "Elite (Nível 5)", "Competitivo (Nível 6)"
                ], index=["Iniciante (Nível 1)", "Básico (Nível 2)", "Intermediário (Nível 3)",
                          "Avançado (Nível 4)", "Elite (Nível 5)", "Competitivo (Nível 6)"].index(cliente['nivel']))
                objetivo = st.selectbox("Objetivo principal", ["Hipertrofia", "Força Máxima", "Potência", "Emagrecimento"],
                                        index=["Hipertrofia", "Força Máxima", "Potência", "Emagrecimento"].index(cliente['objetivo']))
                modalidade = st.selectbox("Modalidade esportiva", [
                    "Geral", "Beach Tennis", "Futebol", "Criança/Adolescente", "Gestante",
                    "Powerlifting", "Fisiculturismo", "Musculação Convencional", "V-Taper", "Bikini"
                ], index=["Geral", "Beach Tennis", "Futebol", "Criança/Adolescente", "Gestante",
                          "Powerlifting", "Fisiculturismo", "Musculação Convencional", "V-Taper", "Bikini"].index(cliente['modalidade']))
                agach = st.number_input("Agachamento (kg)", 0.0, 500.0, cliente['agachamento_1rm'])
                sup = st.number_input("Supino (kg)", 0.0, 500.0, cliente['supino_1rm'])
                terra = st.number_input("Terra (kg)", 0.0, 500.0, cliente['terra_1rm'])
                peg_dir = st.number_input("Força de Pegada Mão Direita (kg)", 0.0, 200.0, cliente['pegada_direita'])
                peg_esq = st.number_input("Força de Pegada Mão Esquerda (kg)", 0.0, 200.0, cliente['pegada_esquerda'])
                if st.form_submit_button("Atualizar"):
                    atualizar_cliente(id_cliente, nome, sexo, idade, peso, altura, perc_gord, nivel, objetivo, modalidade, agach, sup, terra, peg_dir, peg_esq)
                    st.success("Cliente atualizado!")
        with tab2:
            if st.button("Excluir Permanentemente"):
                excluir_cliente(id_cliente)
                st.success("Cliente removido.")
                st.rerun()

elif menu == "Avaliação & Fotos":
    st.header("📸 Fotos Avaliativas e Análise Postural")
    clientes_df = carregar_clientes()
    if clientes_df.empty:
        st.warning("Nenhum cliente cadastrado.")
    else:
        cliente_selecionado = st.selectbox("Selecione o cliente", clientes_df['nome'])
        id_cliente = clientes_df[clientes_df['nome'] == cliente_selecionado]['id'].values[0]
        data_foto = st.date_input("Data da foto", datetime.now())
        col1, col2, col3 = st.columns(3)
        frente = col1.file_uploader("Frente", type=['jpg','jpeg','png'])
        costas = col2.file_uploader("Costas", type=['jpg','jpeg','png'])
        perfil = col3.file_uploader("Perfil", type=['jpg','jpeg','png'])
        if st.button("Salvar Fotos"):
            if not os.path.exists("fotos"):
                os.makedirs("fotos")
            for img, tipo in [(frente, "frente"), (costas, "costas"), (perfil, "perfil")]:
                if img:
                    img_pil = Image.open(img)
                    img_pil.save(f"fotos/{cliente_selecionado}_{data_foto}_{tipo}.png")
            st.success("Fotos salvas! Preencha a avaliação postural abaixo.")
            st.session_state.mostrar_postural = True

        if st.session_state.get('mostrar_postural', False):
            st.subheader("🔍 Avaliação Postural")
            with st.form("form_postural"):
                cabeca = st.selectbox("Cabeça", ["Normal", "Anteriorizada", "Inclinada D", "Inclinada E"])
                ombros = st.selectbox("Ombros", ["Normal", "Protrusos", "Elevado D", "Elevado E", "Desnivelados", "Escápula Alada D", "Escápula Alada E"])
                coluna = st.selectbox("Coluna", ["Normal", "Cifose", "Hiperlordose", "Escoliose"])
                quadril = st.selectbox("Quadril", ["Normal", "Anteroversão", "Retroversão", "Inclinação D", "Inclinação E"])
                joelhos = st.selectbox("Joelhos", ["Normal", "Valgo D", "Valgo E", "Varo D", "Varo E", "Recurvado"])
                pes = st.selectbox("Pés", ["Normal", "Pronado", "Supinado", "Cavo", "Plano"])
                if st.form_submit_button("Salvar Avaliação"):
                    salvar_avaliacao_postural(id_cliente, data_foto.strftime("%Y-%m-%d"), cabeca, ombros, coluna, quadril, joelhos, pes)
                    st.success("Avaliação postural registrada!")
                    st.session_state.mostrar_postural = False
                    st.rerun()

        st.subheader("📊 Comparação de Fotos")
        fotos_df = carregar_fotos(id_cliente)
        if not fotos_df.empty:
            datas = fotos_df['data'].unique()
            if len(datas) >= 2:
                data1 = st.selectbox("Data 1", datas, key="data1")
                data2 = st.selectbox("Data 2", [d for d in datas if d != data1], key="data2")
                if data1 and data2:
                    c1, c2, c3 = st.columns(3)
                    for col, tipo in [(c1, 'frente'), (c2, 'costas'), (c3, 'perfil')]:
                        img1 = fotos_df[(fotos_df['data']==data1)][f'foto_{tipo}'].iloc[0] if not fotos_df[(fotos_df['data']==data1)].empty else None
                        img2 = fotos_df[(fotos_df['data']==data2)][f'foto_{tipo}'].iloc[0] if not fotos_df[(fotos_df['data']==data2)].empty else None
                        if img1:
                            col.image(img1, caption=f"{data1} - {tipo}", use_column_width=True)
                        if img2:
                            col.image(img2, caption=f"{data2} - {tipo}", use_column_width=True)

elif menu == "Geração de Treino":
    st.header("📋 Planilha de Periodização Ondulatória")
    clientes_df = carregar_clientes()
    if clientes_df.empty:
        st.warning("Cadastre um cliente primeiro.")
    else:
        cliente_nome = st.selectbox("Escolha o cliente", clientes_df['nome'])
        id_cliente = clientes_df[clientes_df['nome'] == cliente_nome]['id'].values[0]
        cliente = carregar_cliente(id_cliente)
        if cliente is not None:
            st.write(f"**{cliente['nome']}** | Objetivo: {cliente['objetivo']} | Modalidade: {cliente['modalidade']}")
            semanas = st.slider("Semanas", 4, 12, 12, step=4)
            freq = st.radio("Dias por semana", [2, 3, 4, 5])
            if st.button("Gerar Planilha"):
                df = gerar_planilha(cliente, semanas=semanas, frequencia=freq)
                st.dataframe(df)
                st.markdown(get_table_download_link(df, cliente_nome), unsafe_allow_html=True)

elif menu == "Plano Alimentar":
    st.header("🍽️ Plano Alimentar Personalizado")
    clientes_df = carregar_clientes()
    if clientes_df.empty:
        st.warning("Cadastre um cliente primeiro.")
    else:
        cliente_nome = st.selectbox("Escolha o cliente", clientes_df['nome'])
        id_cliente = clientes_df[clientes_df['nome'] == cliente_nome]['id'].values[0]
        cliente = carregar_cliente(id_cliente)
        if cliente is not None:
            st.write(f"**{cliente['nome']}** | Peso: {cliente['peso']} kg | Altura: {cliente['altura']} cm | % Gordura: {cliente['percentual_gordura']}%")
            fator_atividade = st.selectbox("Nível de atividade física", ["Sedentário", "Pouco ativo", "Moderadamente ativo", "Muito ativo", "Atleta"])
            objetivo_cal = st.selectbox("Meta calórica", ["Manutenção", "Superávit (+300 kcal)", "Déficit (-500 kcal)"])
            if st.button("Gerar Plano Alimentar"):
                tmb = calcular_tmb(cliente['peso'], cliente['altura'], cliente['idade'], cliente['sexo'])
                get = calcular_get(tmb, fator_atividade)
                if "Superávit" in objetivo_cal:
                    meta_cal = get + 300
                elif "Déficit" in objetivo_cal:
                    meta_cal = max(get - 500, 1200)
                else:
                    meta_cal = get

                prot_g, gord_g, carb_g, _ = distribuir_macros(cliente['peso'], cliente['objetivo'], cliente['modalidade'], meta_cal, cliente['sexo'])
                metas = {
                    'tmb': tmb,
                    'get': get,
                    'meta_cal': meta_cal,
                    'prot_g': prot_g,
                    'carb_g': carb_g,
                    'gord_g': gord_g,
                    'cal_prot': prot_g * 4,
                    'cal_carb': carb_g * 4,
                    'cal_gord': gord_g * 9
                }
                refeicoes = montar_refeicoes(meta_cal, prot_g, carb_g, gord_g)

                st.subheader("Distribuição de Macronutrientes")
                st.write(f"Calorias diárias: {meta_cal:.0f} kcal | Proteína: {prot_g}g | Carboidrato: {carb_g}g | Gordura: {gord_g}g")

                for ref in refeicoes:
                    with st.expander(ref['nome']):
                        for item in ref['itens']:
                            st.write(f"- {item[0]} ({item[1]}g): {item[2]} kcal, P:{item[3]}g, C:{item[4]}g, G:{item[5]}g")
                        st.write(f"**Total da refeição:** {ref['total_cal']} kcal, P:{ref['total_prot']}g, C:{ref['total_carb']}g, G:{ref['total_gord']}g")

                output = exportar_plano_alimentar(cliente, refeicoes, metas)
                st.download_button("Baixar Plano em Excel", output, f"dieta_{cliente_nome}.xlsx")

elif menu == "Histórico & Evolução":
    st.header("📈 Histórico do Cliente")
    clientes_df = carregar_clientes()
    if clientes_df.empty:
        st.warning("Nenhum cliente cadastrado.")
    else:
        nome = st.selectbox("Cliente", clientes_df['nome'])
        id_cliente = int(clientes_df[clientes_df['nome']==nome]['id'].values[0])
        cliente = carregar_cliente(id_cliente)
        col1, col2, col3 = st.columns(3)
        col1.metric("Agachamento 1RM", f"{cliente['agachamento_1rm']} kg")
        col2.metric("Supino 1RM", f"{cliente['supino_1rm']} kg")
        col3.metric("Terra 1RM", f"{cliente['terra_1rm']} kg")
        st.subheader("✍️ Registrar Treino Realizado")
        with st.form("registrar_treino"):
            data = st.date_input("Data", datetime.now())
            exercicio = st.text_input("Exercício")
            series = st.number_input("Séries", 1, 20, 3)
            reps = st.number_input("Repetições", 1, 50, 10)
            carga = st.number_input("Carga (kg)", 0.0, 500.0, 60.0)
            obs = st.text_area("Observações")
            if st.form_submit_button("Registrar"):
                salvar_treino(id_cliente, data.strftime("%Y-%m-%d"), exercicio, series, reps, carga, obs)
                st.success("Treino registrado!")
        st.subheader("📊 Evolução de Cargas")
        historico = carregar_historico_treinos(id_cliente)
        if not historico.empty:
            exercicio_filtro = st.selectbox("Exercício", historico['exercicio'].unique())
            df_filtrado = historico[historico['exercicio'] == exercicio_filtro]
            if not df_filtrado.empty:
                fig, ax = plt.subplots()
                ax.plot(pd.to_datetime(df_filtrado['data']), df_filtrado['carga'], marker='o', color=AZUL)
                ax.set_title(f"Progresso - {exercicio_filtro}", color=BRANCO)
                ax.set_xlabel("Data", color=BRANCO)
                ax.set_ylabel("Carga (kg)", color=BRANCO)
                ax.tick_params(colors=BRANCO)
                ax.set_facecolor(PRETO)
                fig.patch.set_facecolor(PRETO)
                st.pyplot(fig)

elif menu == "Personalizar Exercícios":
    st.header("🔧 Personalizar Banco de Exercícios")
    with st.form("add_exercicio"):
        categoria = st.selectbox("Categoria", ["Agachamento", "Supino", "Terra", "Desenvolvimento", "Remada", "Acessórios", "Braços", "Torre Única", "Barra Fixa / Paralela", "Pegada"])
        novo_ex = st.text_input("Nome do novo exercício")
        if st.form_submit_button("Adicionar") and novo_ex:
            st.session_state.EXERCICIOS[categoria].append(novo_ex)
            st.success(f"Exercício '{novo_ex}' adicionado à categoria '{categoria}'.")
            st.rerun()
    st.subheader("Exercícios atuais")
    for cat, exs in st.session_state.EXERCICIOS.items():
        with st.expander(f"{cat} ({len(exs)} exercícios)"):
            for ex in exs:
                col1, col2 = st.columns([4,1])
                col1.write(ex)
                if col2.button("🗑️", key=f"del_{cat}_{ex}"):
                    st.session_state.EXERCICIOS[cat].remove(ex)
                    st.rerun()

st.sidebar.markdown("---")
st.sidebar.markdown("© 2018 Ailson Personal Trainer")